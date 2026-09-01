#!/usr/bin/env python3
"""Build data/bosses.json from a spreadsheet of absorptions and resistances.

    python tools/import_xlsx.py "Elden ring bosses strengths and weaknesses.xlsx"
    python tools/import_xlsx.py <file> --dry-run

Two things about the source data drive most of the code here.

**The columns are inverted relative to what we display.** "Absorption" and
"Resistance" describe how well the boss shrugs damage off, so high
absorption is a reason *not* to bring that damage type. The sheet's labels
are already phrased boss-relative ("WEAK AGAINST" = the boss is weak), so
the mapping follows the label rather than the column name.

**"MEDIUM" is the default, not a middle value.** It appears on 50-70 of
109 rows depending on the column, which makes it the neutral case rather
than a mild resistance.

Existing hand-written entries keep their poise, notes and aliases -- none
of which the sheet has. Where the sheet disagrees with a hand-written
value the sheet wins, but every disagreement is printed, because several
of those entries were marked high confidence and turned out to be wrong.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from erdle.bossdb import (  # noqa: E402
    IMMUNE,
    NORMAL,
    RESISTANT,
    WEAK,
    BossDatabase,
    default_data_path,
)

# --- how the sheet's words map onto our four-point scale ------------------
# Absorption columns. The label is boss-relative, so "WEAK AGAINST" means
# the boss is weak to it, which is what we want to recommend.
ABSORPTION = {
    "VERY WEAK AGAINST": WEAK,
    "WEAKER AGAINST": WEAK,
    "WEAK AGAINST": WEAK,
    "NEUTRAL AGAINST": NORMAL,
    "STRONG AGAINST": RESISTANT,
    "STRONGER AGAINST": RESISTANT,
    "VERY STRONG AGAINST": RESISTANT,
    "VERY STONG AGAINST": RESISTANT,      # typo in the source
}

# Finer ordering, used only to decide which type to name when several are
# equally "resistant" on the coarse scale. Higher means better to bring.
SEVERITY = {
    "VERY WEAK AGAINST": 5,
    "WEAKER AGAINST": 4,
    "WEAK AGAINST": 4,
    "NEUTRAL AGAINST": 3,
    "STRONG AGAINST": 2,
    "STRONGER AGAINST": 1,
    "VERY STRONG AGAINST": 0,
    "VERY STONG AGAINST": 0,
}

# Resistance columns. These are inverted: high resistance means the status
# does not work.
RESISTANCE = {
    "IMMUNE": IMMUNE,
    "HIGH": RESISTANT,
    "MEDIUM": NORMAL,
    "NEUTRAL": NORMAL,
    "LOW": WEAK,
}

COLUMNS = {
    "Standard Absorption": ("damage", "standard"),
    "Slash Absorption": ("damage", "slash"),
    "Strike Absorption": ("damage", "strike"),
    "Thrust Absorption": ("damage", "pierce"),
    "Magic Absorption": ("damage", "magic"),
    "Fire Absorption": ("damage", "fire"),
    "Lightning Absorption": ("damage", "lightning"),
    "Holy Absorption": ("damage", "holy"),
    "Poison Resistance": ("statuses", "poison"),
    "Scarlet Rot Resistance": ("statuses", "rot"),
    "Bleed Resistance": ("statuses", "bleed"),
    "Frostbite Resistance": ("statuses", "frost"),
}

# Misspellings in the source. These matter: the stored name is what OCR
# output is matched against, so a typo costs similarity on every read.
NAME_FIXES = {
    "ADAN, THIEF OF HRE": "Adan, Thief of Fire",
    "SAGUINE NOBLE": "Sanguine Noble",
    "UNCERATED TREE SPIRIT": "Ulcerated Tree Spirit",
    "NIGHT'S CALVALRY": "Night's Cavalry",
    "LORETTA, KNIGHT OF THE HAILGTREE": "Loretta, Knight of the Haligtree",
    "DEMI-HUMAN QUEEN MARGOT": "Demi-Human Queen Margot",
    "ASTEL, NATURALBORN OF THE VOID": "Astel, Naturalborn of the Void",
    # Verified against the wiki roster. Both carry full stats, so the
    # misspelling cost two working bosses: OCR reads the correct name and
    # scores it against a string the game never renders.
    "DRAGONKIN SOLDIER OF NOKESTELLA": "Dragonkin Soldier of Nokstella",
    "DEMI-HUMAN QUEEN GILKA": "Demi-Human Queen Gilika",
}

# Bosses the spreadsheet has no row for at all. Names only -- inventing
# resistances would be worse than admitting there are none, and the panel
# says "no data recorded for this boss" rather than guessing. Detecting
# the fight and naming it is most of the value; the numbers can follow.
EXTRA_BOSSES = [
    "Ancestral Follower",
    "Bloodhound Knight",
    "Bloodhound Knight Floh",
    "Cleanrot Knight Finlay",
    "Crystalian Ringblade",
    "Crystalian Spear",
    "Crystalian Staff",
    "Frenzied Duelist",
    "Godskin Apostle and Godskin Noble",
    "Godskin Duo",
    "Hoarah Loux, Warrior",
    "Necromancer Garris",
    "Nox Monk",
    "Patches",
    "Putrid Crystalian",
    "Putrid Grave Warden Duelist",
    "Vyke, Knight of the Roundtable",
]

# Rows carrying two names. The game shows them at different moments, so
# both need to resolve to the same entry.
SLASH_SPLITS = {
    "MALENIA, BLADE OF MIQUELLA/GODESS OF ROT": (
        "Malenia, Blade of Miquella", ["Malenia, Goddess of Rot"],
    ),
    "MIRANDA THE BLIGHTED BLOOM/BLOSSOM": (
        "Miranda the Blighted Bloom", ["Miranda the Blighted Blossom"],
    ),
    "STRAY MIMIC TEAR/MIMIC TEAR": ("Mimic Tear", ["Stray Mimic Tear"]),
}

# Words the game leaves lowercase inside a title.
PARTICLES = {"of", "the", "and", "in", "de"}

SKIP = {"GODSKIN DUO (CHECK GODSKIN APOSTLE & NOBLE BELOW)"}


def title_case(name: str) -> str:
    """ALL CAPS to the game's rendering, particles left lowercase."""
    words = name.split()
    out = []
    for index, word in enumerate(words):
        lowered = word.lower()
        stripped = lowered.strip(",")
        if index > 0 and stripped in PARTICLES:
            out.append(lowered)
            continue
        # Preserve internal punctuation: O'Neil, Demi-Human, Fia's.
        # A single letter after an apostrophe is a possessive and stays
        # lowercase; a longer fragment is a name and does not.
        parts = _split_keep(lowered)
        rebuilt = []
        for position, part in enumerate(parts):
            if not part.isalpha():
                rebuilt.append(part)
                continue
            possessive = (
                position > 0 and parts[position - 1] == "'" and len(part) == 1
            )
            rebuilt.append(part if possessive else part.capitalize())
        out.append("".join(rebuilt))
    return " ".join(out)


def _split_keep(word: str) -> list[str]:
    parts, current = [], ""
    for char in word:
        if char.isalpha():
            current += char
        else:
            if current:
                parts.append(current)
                current = ""
            parts.append(char)
    if current:
        parts.append(current)
    return parts


def clean_name(raw: str) -> tuple[str, list[str]]:
    """Return the display name and any aliases."""
    raw = raw.strip()
    if raw in SLASH_SPLITS:
        return SLASH_SPLITS[raw]
    if raw in NAME_FIXES:
        return NAME_FIXES[raw], []

    aliases: list[str] = []
    # "CRUCIBLE KNIGHT (CAPITAL)" -- the game shows only "Crucible Knight".
    if "(" in raw:
        base = raw[: raw.index("(")].strip()
        qualifier = raw[raw.index("(") + 1 : raw.rindex(")")].strip()
        name = NAME_FIXES.get(base, title_case(base))
        if qualifier and qualifier.upper() not in {"ALL"}:
            aliases.append(f"{name} ({qualifier.title()})")
        return name, aliases

    return NAME_FIXES.get(raw, title_case(raw)), []


def key_for(name: str) -> str:
    out = []
    for char in name.lower():
        if char.isalnum():
            out.append(char)
        elif out and out[-1] != "_":
            out.append("_")
    return "".join(out).strip("_")


def more_cautious(a: int, b: int) -> int:
    """Merge duplicates by keeping the less encouraging value.

    Two rows for the same enemy should never combine into a stronger
    recommendation than either supports. Under-promising costs a missed
    opportunity; over-promising costs a death.
    """
    return min(a, b)


def load_sheet(path: Path) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        raise SystemExit("needs `pip install openpyxl`")

    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(h).strip() if h else "" for h in rows[0]]

    unknown_values: set[str] = set()
    entries: dict[str, dict] = {}

    for row in rows[1:]:
        if not row or not row[0]:
            continue
        raw = str(row[0]).strip()
        if raw in SKIP or not row[1]:
            continue

        name, aliases = clean_name(raw)
        key = key_for(name)
        entry = entries.setdefault(
            key,
            {
                "name": name, "statuses": {}, "damage": {},
                "severity": {}, "aliases": [], "sources": [],
            },
        )
        for alias in aliases:
            if alias not in entry["aliases"]:
                entry["aliases"].append(alias)
        entry["sources"].append(raw)

        for index, column in enumerate(header):
            target = COLUMNS.get(column)
            if target is None or index >= len(row) or row[index] is None:
                continue
            value = str(row[index]).strip().upper()
            group, field = target
            table = ABSORPTION if group == "damage" else RESISTANCE
            if value not in table:
                unknown_values.add(f"{column}={value}")
                continue
            scored = table[value]
            existing = entry[group].get(field)
            entry[group][field] = (
                scored if existing is None else more_cautious(existing, scored)
            )
            if group == "damage":
                rank = SEVERITY.get(value, 3)
                prior = entry["severity"].get(field)
                entry["severity"][field] = (
                    rank if prior is None else min(prior, rank)
                )

    if unknown_values:
        print("unrecognised cell values (ignored):", file=sys.stderr)
        for value in sorted(unknown_values):
            print(f"  {value}", file=sys.stderr)
    return list(entries.values())


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("spreadsheet")
    parser.add_argument("--out", default=None)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    imported = load_sheet(Path(args.spreadsheet))
    print(f"read {len(imported)} bosses from the spreadsheet")

    existing = BossDatabase.load(default_data_path())
    by_key = {e.key: e for e in existing}
    # Match on the display name too. The sheet's "GODRICK THE GRAFTED"
    # becomes key `godrick_the_grafted`, but the hand-written entry is
    # `godrick` -- same boss. Reusing the existing key keeps notes, poise
    # and every test reference intact, and avoids two entries claiming the
    # same name.
    by_name = {e.name.strip().lower(): e for e in existing}

    conflicts: list[str] = []
    bosses: dict[str, dict] = {}

    for item in imported:
        key = key_for(item["name"])
        prior = by_key.get(key) or by_name.get(item["name"].strip().lower())
        if prior is not None:
            key = prior.key            # keep the established key
        item["key"] = key

        entry: dict = {
            "name": item["name"],
            "statuses": item["statuses"],
            "damage": {k: v for k, v in item["damage"].items() if v != NORMAL},
            "severity": item["severity"],
            "confidence": "sheet",
        }
        if item["aliases"]:
            entry["aliases"] = item["aliases"]

        if prior is not None:
            # Keep what the sheet does not carry.
            if prior.poise is not None:
                entry["poise"] = prior.poise
            if prior.note:
                entry["note"] = prior.note
            if prior.short:
                entry["short"] = prior.short
            merged = list(entry.get("aliases", []))
            for alias in prior.aliases:
                if alias not in merged:
                    merged.append(alias)
            if merged:
                entry["aliases"] = merged

            for field, value in item["statuses"].items():
                was = prior.statuses.get(field)
                if was is not None and was != value:
                    conflicts.append(
                        f"  {prior.name}: {field} was {was} "
                        f"({prior.confidence}), sheet says {value}"
                    )
            for field, value in item["damage"].items():
                was = prior.damage.get(field)
                if was is not None and was != value:
                    conflicts.append(
                        f"  {prior.name}: {field} was {was} "
                        f"({prior.confidence}), sheet says {value}"
                    )

        bosses[key] = entry

    # Bosses the sheet does not mention. Added before the carry-over loop
    # so a from-scratch import produces them too, not just a re-import
    # that happens to have them in the previous file.
    for name in EXTRA_BOSSES:
        key = key_for(name)
        if key in bosses:
            continue
        bosses[key] = {
            "name": name, "statuses": {}, "damage": {}, "severity": {},
            "confidence": "name-only",
        }

    kept = 0
    for entry in existing:
        if entry.key not in bosses:
            payload = {
                "name": entry.name, "statuses": entry.statuses,
                "damage": entry.damage, "confidence": entry.confidence,
            }
            if entry.poise is not None:
                payload["poise"] = entry.poise
            if entry.note:
                payload["note"] = entry.note
            if entry.short:
                payload["short"] = entry.short
            if entry.aliases:
                payload["aliases"] = list(entry.aliases)
            bosses[entry.key] = payload
            kept += 1

    # An alias that equals another entry's real name shadows it: an exact
    # alias hit scores 1.0, so the wrong boss wins. "Beast Clergyman" was
    # an alias on Maliketh (his second phase) until the sheet supplied a
    # dedicated entry for the first phase.
    real_names = {e["name"].strip().lower() for e in bosses.values()}
    shadowed = []
    for entry in bosses.values():
        kept = [a for a in entry.get("aliases", [])
                if a.strip().lower() not in real_names]
        dropped = [a for a in entry.get("aliases", []) if a not in kept]
        for alias in dropped:
            shadowed.append(f"  {entry['name']}: dropped alias {alias!r}")
        if kept:
            entry["aliases"] = kept
        else:
            entry.pop("aliases", None)
    if shadowed:
        print(f"\n{len(shadowed)} aliases removed for shadowing a real name:")
        for line in shadowed:
            print(line)

    if conflicts:
        print(f"\n{len(conflicts)} disagreements with hand-written values:")
        for line in sorted(set(conflicts)):
            print(line)
        print("  (the sheet wins; these hand-written values were guesses)")

    print(f"\n{len(bosses)} bosses total ({kept} hand-written kept as-is)")

    payload = {
        "schema_version": 1,
        "source": "spreadsheet",
        "_warning": (
            "Absorption and resistance values imported from a community "
            "spreadsheet via tools/import_xlsx.py. Poise values and notes "
            "are hand-written and still approximate."
        ),
        "scale": {"0": "immune", "1": "resistant", "2": "normal", "3": "weak"},
        "bosses": dict(sorted(bosses.items())),
    }

    if args.dry_run:
        print("\n(dry run -- nothing written)")
        return 0

    target = Path(args.out) if args.out else default_data_path()
    target.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    print(f"wrote {target}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
