#!/usr/bin/env python3
"""Fill data/bosses.json from the community PvE stats workbook.

    python tools/import_pve_stats.py --check
    python tools/import_pve_stats.py

The workbook carries what nothing else did in one place: the NpcParam row
ID *and* a readable boss name, plus damage negation, status resistances
and poise for 230 boss rows.

Two things make it better than the regulation.bin route, which is why it
supersedes it:

* **Damage negation is the stat that decides a fight.** `NpcParam`'s
  `DamageCutRate` is a second, subtler multiplier -- Crystalians read 0.9
  strike against 0.65 slash there, but their *negation* is 10 against 94.
  Same conclusion, hugely different signal, and the negation figure is the
  one a player feels.
* **Negative values exist.** A negation of -20 means the boss takes 20%
  *more*, which is a real weakness the four-bucket scale had no way to
  express and the old spreadsheet never recorded.

Column layout is read from the header row rather than hard-coded offsets,
because the poise block sits at columns 40-43 and it is easy to grab
"Incoming Mult" or "Regen Delay" by accident -- which happened, and gave
Godrick a poise of 8 instead of 105.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from erdle.bossdb import IMMUNE, NORMAL, RESISTANT, WEAK  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
#: Research inputs, not runtime data. They live under tools/ because
#: `data/` is what the exe bundles, and shipping a 65 KB spreadsheet that
#: nothing reads at run time only invites the question of what reads it.
SOURCES = ROOT / "tools" / "sources"

WORKBOOK = SOURCES / "ER - PvE Health_Defense_DmgNeg_Resistances.xlsx"
#: Bosses the workbook omits, hand-transcribed in the same units. Kept as
#: a separate file so re-downloading the workbook cannot silently drop
#: them, and so adding one more is a line of CSV rather than a code edit.
EXTRA_ROWS = SOURCES / "extra_pve_rows.csv"
DATABASE = ROOT / "data" / "bosses.json"

#: Damage-negation columns, by the header text above them.
NEGATION = {
    "standard": 17, "strike": 18, "slash": 19, "pierce": 20,
    "magic": 21, "fire": 22, "lightning": 23, "holy": 24,
}
RESISTANCE = {
    "poison": 26, "rot": 27, "bleed": 28,
    "frost": 29, "sleep": 30, "madness": 31,
}
POISE_EFFECTIVE = 42

#: Names the workbook spells differently. Kept short and explicit: a fuzzy
#: matcher here would quietly hand a named variant its generic's numbers,
#: which is the failure this data has already had once.
ALIASES = {
    "bloodhound_knight_darriwil": "Bloodhound Knight Darriwill",
    "fell_twins": "Fell Twin (Axe)",
    # The Swordstress and the Priest carry identical profiles, so either
    # describes the pair. Verified, not assumed -- see the test.
    "nox_swordstress": "Nox Swordstress",
    "nox_priest": "Nox Priest",
}


def normalise(text: str) -> str:
    return re.sub(r"[^a-z0-9]", "", text.lower())


def bucket_for(negation: float) -> int:
    """Negation is a percentage of damage removed, so it can go negative."""
    if negation >= 100:
        return IMMUNE
    if negation < 0:
        return WEAK
    if negation == 0:
        return NORMAL
    return RESISTANT


def severity_for(negation: float) -> int:
    """The six-point scale, higher meaning better to bring."""
    if negation <= -20:
        return 5
    if negation < 0:
        return 4
    if negation == 0:
        return 3
    if negation <= 20:
        return 2
    if negation <= 50:
        return 1
    return 0


#: Below this, a status procs easily on anyone. From the distribution
#: across all 230 rows, whose deciles run 229 / 318 / 351 / 539 / 689.
WEAK_ABSOLUTE = 260
NORMAL_CEILING = 560
#: A status also counts as weak when it is far below the rest of *this*
#: boss's resistances. Absolute cuts alone reported Malenia -- the most
#: fought boss in the game -- as having no notable status at all, when
#: frost at 306 against her poison at 1481 is the whole strategy.
RELATIVE_WEAK = 0.6


def status_for(value, peers=()) -> int | None:
    """Resistance is the build-up needed to proc: lower is easier.

    `peers` are the same boss's other resistances. Judging only against
    them would call Godrick weak to everything, since all five of his sit
    at 318; judging only absolutely misses Malenia. A status is weak if
    either test says so.
    """
    if isinstance(value, str):
        return IMMUNE if value.strip().lower().startswith("immune") else None
    if value is None:
        return None
    if value <= WEAK_ABSOLUTE:
        return WEAK

    numbers = sorted(v for v in peers if isinstance(v, (int, float)))
    if numbers:
        middle = numbers[len(numbers) // 2]
        if value <= middle * RELATIVE_WEAK:
            return WEAK

    return NORMAL if value <= NORMAL_CEILING else RESISTANT


def read_workbook(path: Path) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        raise SystemExit("needs `pip install openpyxl`")

    sheet = openpyxl.load_workbook(path, data_only=True)["Sheet1"]
    for field, column in (("Base", 40), ("Effective", POISE_EFFECTIVE)):
        header = sheet.cell(2, column).value
        if header != field:
            raise SystemExit(
                f"column {column} is {header!r}, expected {field!r} -- "
                "the workbook layout has changed"
            )

    rows = []
    for index in range(3, sheet.max_row + 1):
        raw = sheet.cell(index, 2).value
        if not raw:
            continue
        rows.append({
            "name": re.sub(r"\s*\[Boss\]\s*$", "", str(raw).strip()),
            "id": sheet.cell(index, 3).value,
            "negation": {k: sheet.cell(index, c).value
                         for k, c in NEGATION.items()},
            "resistance": {k: sheet.cell(index, c).value
                           for k, c in RESISTANCE.items()},
            "poise": sheet.cell(index, POISE_EFFECTIVE).value,
        })
    return rows



def read_extra_rows(path: Path) -> list[dict]:
    """Hand-transcribed rows in the workbook's own units.

    Read by field name rather than position: Fextralife lists Standard,
    Slash, Strike, Pierce while the workbook column order puts Strike
    before Slash, and transcribing by eye down two differently-ordered
    lists is exactly how a boss ends up with its slash and strike swapped.
    """
    import csv

    if not path.exists():
        return []

    def number(cell: str):
        cell = (cell or "").strip()
        if not cell:
            return None
        if cell.lower().startswith("immune"):
            return "Immune"
        return float(cell)

    rows = []
    with path.open(encoding="utf-8-sig", newline="") as handle:
        lines = (line for line in handle if not line.lstrip().startswith("#"))
        for record in csv.DictReader(lines):
            name = (record.get("name") or "").strip()
            if not name:
                continue
            poise = (record.get("poise") or "").strip()
            rows.append({
                "name": name,
                "id": f"hand-entered from {path.name}",
                "negation": {k: number(record.get(k)) for k in NEGATION},
                "resistance": {k: number(record.get(k)) for k in RESISTANCE},
                "poise": float(poise) if poise else None,
            })
    return rows


def index_rows(rows: list[dict]) -> tuple[dict, dict]:
    """Two lookups: with the "(Sickle)" variant, and without."""
    exact, base = defaultdict(list), defaultdict(list)
    for row in rows:
        exact[normalise(row["name"])].append(row)
        stripped = re.sub(r"\s*\([^)]*\)", "", row["name"])
        base[normalise(stripped)].append(row)
    return exact, base


def damage_profile(row: dict) -> tuple:
    """What this row would contribute to damage, not its raw numbers.

    Placements of the same boss carry slightly different figures -- 530
    here, 556 there -- which are the same answer once bucketed. Comparing
    raw values called thirty bosses ambiguous when every one of them
    agreed on the only thing that gets written.
    """
    return tuple(
        (bucket_for(v), severity_for(v)) if isinstance(v, (int, float)) else None
        for v in row["negation"].values()
    )


def status_profile(row: dict) -> tuple:
    return tuple(status_for(v, row["resistance"].values())
                 for v in row["resistance"].values())



def pick_status_row(candidates: list[dict]) -> dict | None:
    """Which placement's resistances to believe when they differ.

    Two things cause the disagreement, and they need different answers:

    * A variant that is immune to everything is the non-fightable one --
      a cutscene or invulnerable phase. Godfrey has one, and keeping it
      in the running is what made him read as immune to four statuses.
    * The rest differ only by level scaling: the Bell Bearing Hunter's
      four placements run 316 / 329 / 337 / 383. The middle one is the
      least wrong answer, and picking the lowest would over-claim.
    """
    fightable = [r for r in candidates
                 if any(isinstance(v, (int, float))
                        for v in r["resistance"].values())]
    if not fightable:
        return candidates[0] if candidates else None

    def middle(row):
        numbers = sorted(v for v in row["resistance"].values()
                         if isinstance(v, (int, float)))
        return numbers[len(numbers) // 2] if numbers else 0

    return sorted(fightable, key=middle)[len(fightable) // 2]


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--check", action="store_true")
    parser.add_argument("--workbook", type=Path, default=WORKBOOK)
    args = parser.parse_args()

    if not args.workbook.exists():
        print(f"no workbook at {args.workbook}", file=sys.stderr)
        return 1

    rows = read_workbook(args.workbook) + read_extra_rows(EXTRA_ROWS)
    exact, base = index_rows(rows)
    doc = json.loads(DATABASE.read_text(encoding="utf-8"))
    bosses = doc["bosses"]

    applied, ambiguous, unmatched, partial = 0, [], [], []
    for key, entry in bosses.items():
        name = entry["name"]
        candidates = (
            exact.get(normalise(ALIASES.get(key, ""))) if key in ALIASES else None
        ) or exact.get(normalise(name)) or base.get(normalise(name))
        if not candidates and "," in name:
            head = normalise(name.split(",")[0])
            if len(head) >= 5:
                candidates = exact.get(head) or base.get(head)
        if not candidates:
            unmatched.append(name)
            continue

        # Several rows can share a name, one per placement. The two halves
        # are judged separately: Bell Bearing Hunter's four placements
        # carry identical negation and different status resistances, and
        # discarding the damage data over that threw away the half that
        # actually decides a fight.
        row = candidates[0]
        damage_agrees = len({damage_profile(r) for r in candidates}) == 1
        status_agrees = len({status_profile(r) for r in candidates}) == 1
        if not damage_agrees and not status_agrees:
            ambiguous.append(name)
            continue

        damage, severity, statuses = {}, {}, {}
        if damage_agrees:
            for field, value in row["negation"].items():
                if not isinstance(value, (int, float)):
                    continue
                got = bucket_for(value)
                severity[field] = severity_for(value)
                if got != NORMAL:
                    damage[field] = got
        else:
            damage = dict(entry.get("damage", {}))
            severity = dict(entry.get("severity", {}))
            partial.append(f"{name} (damage only from the old source)")

        status_row = row if status_agrees else pick_status_row(candidates)
        if status_row is not None:
            for field, value in status_row["resistance"].items():
                got = status_for(value, status_row["resistance"].values())
                if got is not None:
                    statuses[field] = got
        if not status_agrees:
            partial.append(f"{name} (statuses vary by placement; middle taken)")

        if not severity and not statuses:
            continue

        applied += 1
        if args.check:
            continue

        entry["damage"] = damage
        entry["severity"] = severity
        entry["statuses"] = statuses
        poise = row["poise"]
        if isinstance(poise, (int, float)) and poise >= 0:
            entry["poise"] = int(round(poise))
        else:
            # "∞" for Rennala phase one: she cannot be staggered, so there
            # is no threshold to report.
            entry.pop("poise", None)
        entry["confidence"] = "pve-sheet"
        entry["note"] = (f"PvE workbook, NpcParam row {row['id']}"
                         if isinstance(row["id"], int) else row["id"])

    print(f"{applied} bosses filled from the workbook")
    if partial:
        print(f"partly filled ({len(partial)}): {'; '.join(sorted(set(partial)))}")
    if ambiguous:
        print(f"ambiguous, left alone ({len(ambiguous)}): {', '.join(sorted(ambiguous))}")
    if unmatched:
        print(f"not in the workbook ({len(unmatched)}): {', '.join(sorted(unmatched))}")
    if args.check:
        print("(--check: nothing written)")
        return 0

    doc["bosses"] = dict(sorted(bosses.items()))
    doc["source"] = "PvE stats workbook + regulation.bin + spreadsheet"
    DATABASE.write_text(
        json.dumps(doc, indent=1, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    print(f"written to {DATABASE}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
